#!/usr/bin/env python3
"""
AR Aging Report Sync - SharePoint to Supabase
Syncs Petra Brands AR aging data daily
"""

import os
import sys
import re
from datetime import datetime, date
from pathlib import Path
import requests
from dotenv import load_dotenv
from azure.identity import ClientSecretCredential
import openpyxl
from supabase import create_client, Client

# Load environment variables
load_dotenv(Path(__file__).parent.parent / '.env')

# Configuration
CONFIG = {
    'tenant_id': os.getenv('AZURE_TENANT_ID'),
    'client_id': os.getenv('AZURE_CLIENT_ID'),
    'client_secret': os.getenv('AZURE_CLIENT_SECRET'),
    'supabase_url': os.getenv('SUPABASE_URL'),
    'supabase_key': os.getenv('SUPABASE_KEY'),
    'sharepoint_site': 'swiftstartagency.sharepoint.com',
    'site_path': '/sites/PetraBrands',
    'file_name': 'HOP Retail Sales & PO Tracker - 2025-2026 - Sales Copy.xlsx',
    'log_file': Path(__file__).parent.parent / 'logs' / 'aging_sync.log',
}

def log(message, level='INFO'):
    """Log with timestamp"""
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    log_msg = f"[{timestamp}] [{level}] {message}"
    print(log_msg)

    CONFIG['log_file'].parent.mkdir(exist_ok=True)
    with open(CONFIG['log_file'], 'a') as f:
        f.write(log_msg + '\n')

def get_access_token():
    """Get Microsoft Graph API access token"""
    credential = ClientSecretCredential(
        CONFIG['tenant_id'],
        CONFIG['client_id'],
        CONFIG['client_secret']
    )
    token = credential.get_token('https://graph.microsoft.com/.default')
    return token.token

def get_site_id(access_token):
    """Get SharePoint site ID"""
    url = f"https://graph.microsoft.com/v1.0/sites/{CONFIG['sharepoint_site']}:{CONFIG['site_path']}"
    headers = {'Authorization': f'Bearer {access_token}'}

    response = requests.get(url, headers=headers)
    response.raise_for_status()
    return response.json()['id']

def find_file(access_token, site_id):
    """Find the aging report file in SharePoint"""
    drives_url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drives"
    headers = {'Authorization': f'Bearer {access_token}'}

    drives_response = requests.get(drives_url, headers=headers)
    drives_response.raise_for_status()
    drives = drives_response.json().get('value', [])

    for drive in drives:
        drive_id = drive['id']
        search_url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/root/search(q='{CONFIG['file_name']}')"

        search_response = requests.get(search_url, headers=headers)
        search_response.raise_for_status()
        items = search_response.json().get('value', [])

        for item in items:
            if CONFIG['file_name'].lower() in item.get('name', '').lower():
                return drive_id, item['id']

    raise FileNotFoundError(f"File not found: {CONFIG['file_name']}")

def download_file(access_token, drive_id, item_id):
    """Download Excel file from SharePoint"""
    content_url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{item_id}/content"
    headers = {'Authorization': f'Bearer {access_token}'}

    response = requests.get(content_url, headers=headers, allow_redirects=True)
    response.raise_for_status()

    temp_file = Path('/tmp/aging_report.xlsx')
    with open(temp_file, 'wb') as f:
        f.write(response.content)

    return temp_file

def normalize_brand(brand):
    """Normalize brand names"""
    brand = str(brand).strip().upper()
    mapping = {
        'FOMIN': 'Fomin',
        'HOP': 'House of Party',
        'ROOFUS': 'Roofus',
        'EVERYMOOD': 'EveryMood',
        'LUNA': 'Luna Naturals'
    }
    return mapping.get(brand, brand.title())

def parse_excel(file_path):
    """Parse aging report Excel file"""
    wb = openpyxl.load_workbook(file_path, data_only=True)

    # Parse Detailed Invoices sheet
    invoices = []
    if 'Detailed Invoices' in wb.sheetnames:
        ws = wb['Detailed Invoices']
        headers = [cell.value for cell in ws[1]]

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:  # Skip empty rows
                continue

            invoice = {
                'brand': normalize_brand(row[0]) if row[0] else None,
                'retailer': str(row[1]).strip() if row[1] else None,
                'po_number': str(row[2]).strip() if row[2] else None,
                'invoice_number': str(row[3]).strip().replace('\n', '') if row[3] else None,
                'invoice_date': row[4] if isinstance(row[4], date) else None,
                'due_date': row[5] if isinstance(row[5], date) else None,
                'invoice_amount': float(row[6]) if row[6] else 0,
                'received': float(row[7]) if row[7] else 0,
                'outstanding': float(row[8]) if row[8] else 0,
                'days_overdue': int(row[9]) if row[9] else 0,
                'aging_bucket': str(row[10]).strip() if row[10] else 'Not Yet Due',
                'status': str(row[11]).strip() if row[11] else 'Not Due',
                'delay_remarks': str(row[12]).strip() if row[12] else None,
                'report_date': date.today(),
            }

            if invoice['invoice_number']:
                invoices.append(invoice)

    log(f"Parsed {len(invoices)} invoices from Excel")

    # Parse Aging Report summary (optional, for now focus on detailed invoices)
    summaries = []

    return invoices, summaries

def sync_to_supabase(invoices, summaries):
    """Upsert data to Supabase"""
    supabase: Client = create_client(CONFIG['supabase_url'], CONFIG['supabase_key'])

    # Sync invoices (batch upsert)
    log(f"Syncing {len(invoices)} invoices to Supabase...")
    batch_size = 100
    upserted = 0

    for i in range(0, len(invoices), batch_size):
        batch = invoices[i:i + batch_size]
        try:
            result = supabase.table('retail_aging_invoices').upsert(
                batch,
                on_conflict='brand,invoice_number'
            ).execute()
            upserted += len(batch)
            log(f"Batch {i//batch_size + 1}: {upserted}/{len(invoices)}")
        except Exception as e:
            log(f"Error upserting batch {i//batch_size + 1}: {str(e)}", 'ERROR')

    # Create daily snapshots
    log("Creating daily snapshots...")
    snapshot_data = []

    for brand in set(inv['brand'] for inv in invoices if inv['brand']):
        brand_invoices = [inv for inv in invoices if inv['brand'] == brand]
        total_outstanding = sum(inv['outstanding'] for inv in brand_invoices)
        not_yet_due = sum(inv['outstanding'] for inv in brand_invoices if inv['aging_bucket'] == 'Not Yet Due')
        overdue = total_outstanding - not_yet_due

        snapshot = {
            'snapshot_date': date.today(),
            'brand': brand,
            'total_outstanding': total_outstanding,
            'not_yet_due': not_yet_due,
            'overdue_total': overdue,
            'invoice_count': len(brand_invoices),
            'overdue_count': sum(1 for inv in brand_invoices if inv['status'] == 'Due'),
        }
        snapshot_data.append(snapshot)

    try:
        supabase.table('retail_aging_snapshots').upsert(
            snapshot_data,
            on_conflict='snapshot_date,brand'
        ).execute()
        log(f"Created {len(snapshot_data)} snapshots")
    except Exception as e:
        log(f"Error creating snapshots: {str(e)}", 'ERROR')

    return upserted

def main():
    """Main sync function"""
    log('=' * 60)
    log('Starting AR Aging Report Sync')
    log('=' * 60)

    try:
        # Validate config
        if not all([CONFIG['tenant_id'], CONFIG['client_id'], CONFIG['client_secret']]):
            raise ValueError("Missing Azure credentials in .env")
        if not all([CONFIG['supabase_url'], CONFIG['supabase_key']]):
            raise ValueError("Missing Supabase credentials in .env")

        # Get access token
        log("Authenticating with Microsoft Graph...")
        access_token = get_access_token()

        # Get site ID
        log("Getting SharePoint site ID...")
        site_id = get_site_id(access_token)
        log(f"Site ID: {site_id}")

        # Find file
        log(f"Searching for file: {CONFIG['file_name']}")
        drive_id, item_id = find_file(access_token, site_id)
        log(f"Found file (Drive: {drive_id}, Item: {item_id})")

        # Download file
        log("Downloading Excel file...")
        file_path = download_file(access_token, drive_id, item_id)
        log(f"Downloaded to {file_path}")

        # Parse Excel
        log("Parsing Excel data...")
        invoices, summaries = parse_excel(file_path)

        # Sync to Supabase
        log("Syncing to Supabase...")
        upserted = sync_to_supabase(invoices, summaries)

        # Clean up
        file_path.unlink()

        log('=' * 60)
        log('Sync Complete')
        log(f"  Invoices synced: {upserted}")
        log(f"  Brands: {len(set(inv['brand'] for inv in invoices if inv['brand']))}")
        log('=' * 60)

    except Exception as e:
        log(f"Sync failed: {str(e)}", 'ERROR')
        import traceback
        log(traceback.format_exc(), 'ERROR')
        sys.exit(1)

if __name__ == '__main__':
    main()
